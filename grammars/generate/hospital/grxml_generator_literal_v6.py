import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
import re
import yaml                     # NEW ←─────────────
 
 # ------------------------------------------------------------------ #
 #  Load spreadsheets
 # ------------------------------------------------------------------ #
data_wb  = openpyxl.load_workbook("Phone-NEW customer copy May 13.xlsx")
data_ws  = data_wb.active
alias_wb = openpyxl.load_workbook("departments.xlsx")
alias_ws = alias_wb.active

# ──────────────────────────────────────────────────────────────
#  NEW – normalize the SPOKEN text (utterances)
#      • "Sr."  →  "senior"
#      • "/"    →  " "
#      • collapse multiple spaces
# ──────────────────────────────────────────────────────────────
def normalize_spoken(text: str) -> str:
    """
    • “Sr.” or “Sr”  →  “senior”   (dot removed)
    • “/”            →  “ ”
    • collapse duplicate whitespace
    """
    if not text:
        return ""

    # 1. slash → space
    txt = text.replace("/", " ")

    # 2. Sr / Sr.  → senior   (match whole word, optional dot, consume the dot)
    #    (?=\s|$)  ensures the abbreviation ends either at a space or end-of-string
    txt = re.sub(r'\bsr\.?(?=\s|$)', 'Senior', txt, flags=re.IGNORECASE)

    # 3. collapse runs of whitespace
    txt = re.sub(r'\s+', ' ', txt).strip()

    return txt


# ------------------------------------------------------------------ #
#  Load staff pronunciation variants        NEW ←────────────────────
#    YAML format:                            (see sample in prompt)
#      “Full Name”:
#        - [variants for first component]
#        - [variants for second …]
#        - …
#  We keep everything *already* in the spoken/normalized form so that
#  later grammar generation is trivial.
# ------------------------------------------------------------------ #
try:
    with open("staff_names_expanded.yaml", "r", encoding="utf-8") as fh:
        _raw_name_variants: dict[str, list[list[str]]] = yaml.safe_load(fh) or {}
except FileNotFoundError:
    _raw_name_variants = {}

def _norm(txt: str) -> str:
    return normalize_spoken(txt or "")

# {normalized full name → list[list[normalized variants]]}
name_variants: dict[str, list[list[str]]] = {
    _norm(full): [[_norm(v) for v in comp] for comp in comps]
     for full, comps in _raw_name_variants.items()
}
# ------------------------------------------------------------------ #
 

# ------------------------------------------------------------------ #
#  Helpers
# ------------------------------------------------------------------ #
grammar = ET.Element(
    "grammar",
    {
        "xmlns":      "http://www.w3.org/2001/06/grammar",
        "xml:lang":   "en-US",
        "mode":       "voice",
        "version":    "1.0",
        "root":       "primary_rule",
        "tag-format": "semantics/1.0-literals",
    },
)

def sanitize(text: str) -> str:
    """
    Convert an arbitrary name (“Brian Smith, Sr.”, “IT/HR”, etc.)
    into a rule-id that meets SRGS XML requirements:
      • a–z, 0–9, underscore, hyphen only
      • must start with a letter
    Extra punctuation collapses to a single “_”.
    """
    if text is None:
        text = ""

    s = text.lower().strip()

    # replace a few common separators up-front
    s = s.replace("&", " and ")

    # convert every run of non-alphanumerics to "_"
    s = re.sub(r"[^a-z0-9]+", "_", s)

    # collapse multiple underscores and trim edges
    s = re.sub(r"_+", "_", s).strip("_")

    # ensure the first char is a letter (SRGS requirement)
    if not s or not s[0].isalpha():
        s = f"r_{s}"

    return s


def clean_tag_part(text: str) -> str:
    """
    For literal tags only:
      • replace '.' and '/' with spaces
      • collapse multiple spaces to one
      • trim
    """
    if text is None:
        return ""
    # dot & slash → space
    txt = text.replace(".", " ").replace("/", " ")
    # collapse runs of whitespace
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()

def create_rule(rule_id: str, scope: str | None = None) -> ET.Element:
    rule = ET.Element("rule", id=sanitize(rule_id))
    if scope:
        rule.set("scope", scope)
    return rule

def create_item(text: str, repeat: str | None = None) -> ET.Element:
    itm = ET.Element("item")
    if repeat:
        itm.set("repeat", repeat)
    itm.text = text.strip()
    return itm

def create_ruleref(rule_id: str) -> ET.Element:
    return ET.Element("ruleref", uri=f"#{sanitize(rule_id)}")

def literal_tag(text: str) -> ET.Element:
    tg = ET.Element("tag"); tg.text = text
    return tg

def add_choice(container: ET.Element, phrases: list[str], repeat: str | None = None):
    """
    Append either:
        <item>phrase</item>
    or
        <one-of><item>p1</item><item>p2</item>…</one-of>
    inside the given container.
    """
    if len(phrases) == 1:
        container.append(create_item(phrases[0], repeat))
    else:
        one_of = ET.Element("one-of")
        if repeat:
            one_of.set("repeat", repeat)
        for p in phrases:
            one_of.append(create_item(p))
        container.append(one_of)

def wrap_rule(rule: ET.Element, spoken_children: list[ET.Element], tag_text: str):
    """
    Build the mandatory outer <item> that contains both spoken elements
    and the literal <tag>, then attach it to the rule.
    """
    wrapper = ET.Element("item")
    for child in spoken_children:
        wrapper.append(child)
    wrapper.append(literal_tag(tag_text))
    rule.append(wrapper)

# ------------------------------------------------------------------ #
#  Build alias map  {canonical → [aliases…]}
# ------------------------------------------------------------------ #
aliases_by_dept: dict[str, list[str]] = {}

for col in alias_ws.iter_cols(1, alias_ws.max_column):
    # skip empty columns or columns whose 2nd cell is empty
    if not col or not col[1].value:
        continue

    canonical = normalize_spoken(str(col[0].value).strip())          # row 2
    raw_aliases = [
        normalize_spoken(str(c.value).strip().replace("&", " and "))
        for c in col
        if c.value and str(c.value).strip()
    ]

    # drop the canonical name (and duplicates) from the list
    aliases = [a for a in raw_aliases if a != canonical]

    aliases_by_dept[canonical] = aliases        # may be empty list
    print(f"Loaded aliases for {canonical}: {aliases}")  # debug output

## print aliases loaded from spredsheet
print("Aliases by department:")
for dept, aliases in aliases_by_dept.items():     # debug output
    print(f"  {dept}: {aliases}")                                             


# ------------------------------------------------------------------ #
#  First pass – gather info
# ------------------------------------------------------------------ #
dept_top_dn: dict[str, str] = {}          # first phone per dept
people_rows: list[tuple]   = []           # (first, last, dept, dn)

for i in range(2, data_ws.max_row + 1):   # row 1 = header
    dn          = str(data_ws.cell(i, 2).value or "").strip()    # B
    dept_cell   = str(data_ws.cell(i, 4).value or "").strip()    # D
    first_name  = str(data_ws.cell(i, 5).value or "").strip()    # E
    last_name   = str(data_ws.cell(i, 6).value or "").strip()    # F
    is_person   = str(data_ws.cell(i, 7).value or "").strip().lower() == "yes"  # G

    if dept_cell and dn and dept_cell not in dept_top_dn:
        dept_top_dn[dept_cell] = dn        # first phone = dept number

    if is_person and first_name and last_name and dn:
        people_rows.append((first_name, last_name, dept_cell, dn))

# ------------------------------------------------------------------ #
#  Create rules
# ------------------------------------------------------------------ #
primary_oneof = ET.Element("one-of")

# ── inside the “Department-only rules” loop ─────────────────────────
for dept, dn in dept_top_dn.items():
    canonical = normalize_spoken(dept)          # spoken & rule-id base
    print(f"Processing department: {canonical} (dn: {dn})")  # debug output

    aliases   = aliases_by_dept.get(canonical, [])   # already excludes canonical
    print(f"  Aliases: {aliases}")  # debug output

    # Build the spoken choices
    phrases = [canonical] + aliases                 # canonical first
    tmp = ET.Element("dummy")
    add_choice(tmp, phrases)                        # one-of if >1
    spoken_parts = list(tmp)

    # 4. wrap + append
    dept_rule = create_rule(dept)
    wrap_rule(
        dept_rule,
        spoken_parts,
        f"department {clean_tag_part(dept)} number {clean_tag_part(dn)}"
    )
    grammar.append(dept_rule)

    pr_item = ET.Element("item")
    pr_item.append(create_ruleref(dept))
    primary_oneof.append(pr_item)

# --- Person rules -------------------------------------------------- #
created_name_only    : set[str] = set()
created_name_and_dept: set[str] = set()

# ------------------------------------------------------------------ #
#  Name-variant helper                    NEW ←──────────────────────
# ------------------------------------------------------------------ #
def build_name_spoken_parts(full_name: str) -> list[ET.Element]:
    """
    Returns a list of <item>/<one-of> elements that together pronounce `full_name`.

    • If we have variants in `name_variants`, we create a <one-of> for each
      component (first, middle, last …).  
      e.g.  Jennifer / Jenifer / Jenniffer   +   Galveias / Galveas
    • Otherwise we fall back to the simple, normalized full name.
    """
    norm_full = normalize_spoken(full_name)
    comps = name_variants.get(norm_full)

    if not comps:
        return [create_item(norm_full)]

    spoken_parts: list[ET.Element] = []
    for comp_variants in comps:
        # Deduplicate while keeping order
        seen = set()
        variants = [v for v in comp_variants if not (v in seen or seen.add(v))]
        tmp = ET.Element("dummy")
        add_choice(tmp, variants)            # <one-of> if >1
        spoken_parts.extend(tmp)
    return spoken_parts


for first, last, dept, dn in people_rows:
    full_name = f"{first} {last}".strip()

    # (a) name-only
    if full_name not in created_name_only:
        name_rule = create_rule(full_name)
        spoken_name = normalize_spoken(full_name)
        wrap_rule(
            name_rule,
            build_name_spoken_parts(full_name),
            f"{clean_tag_part(full_name)} number {clean_tag_part(dn)}"
        )
        grammar.append(name_rule)

        pr_item = ET.Element("item"); pr_item.append(create_ruleref(full_name))
        primary_oneof.append(pr_item)

        created_name_only.add(full_name)

    # (b) name + department
    combo_id = f"{full_name} {dept}"
# ── Person + Department rules ──────────────────────────────────────
    if combo_id not in created_name_and_dept:
        canonical_dept = normalize_spoken(dept)                 # canonical spoken
        alias_list     = aliases_by_dept.get(canonical_dept, [])  # excludes canonical
        # build unique list with canonical first
        phrases, seen = [], set()
        for p in [canonical_dept] + alias_list:
            if p not in seen:
                phrases.append(p)
                seen.add(p)

        combo_rule = create_rule(combo_id)

        # spoken_parts = [
        #     create_item(normalize_spoken(full_name)),
        #     create_item("in", repeat="0-1"),
        # ]
        spoken_parts = build_name_spoken_parts(full_name)
        spoken_parts.append(create_item("in", repeat="0-1"))

        tmp_container = ET.Element("dummy")
        add_choice(tmp_container, phrases)                      # <one-of> if len>1
        spoken_parts.extend(tmp_container)

        wrap_rule(
            combo_rule,
            spoken_parts,
            f"{clean_tag_part(full_name)} in {clean_tag_part(dept)} "
            f"number {clean_tag_part(dn)}"
        )
        grammar.append(combo_rule)

        pr_item = ET.Element("item")
        pr_item.append(create_ruleref(combo_id))
        primary_oneof.append(pr_item)

        created_name_and_dept.add(combo_id)

# ------------------------------------------------------------------ #
#  Root rule
# ------------------------------------------------------------------ #
primary_rule = create_rule("primary_rule", scope="public")
primary_rule.append(primary_oneof)
grammar.append(primary_rule)

# ------------------------------------------------------------------ #
#  Write output
# ------------------------------------------------------------------ #
with open("Smith_Falls_7.grxml", "wb") as fh:
    pretty = minidom.parseString(ET.tostring(grammar, encoding="utf-8"))
    fh.write(pretty.toprettyxml(indent="  ", encoding="UTF-8"))
