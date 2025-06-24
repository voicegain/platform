import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl

# ------------------------------------------------------------------ #
#  Load data
# ------------------------------------------------------------------ #
departments_ws = openpyxl.load_workbook("departments.xlsx").active
data_ws        = openpyxl.load_workbook("Phone-NEW customer copy May 13.xlsx").active

with open("staff_names.txt", "r", encoding="utf-8") as fh:
    staff_names = {n.lower() for n in fh.read().splitlines()}

# ------------------------------------------------------------------ #
#  Helpers
# ------------------------------------------------------------------ #
grammar = ET.Element(
    "grammar",
    {
        "xmlns":      "http://www.w3.org/2001/06/grammar",
        "xml:lang":   "en-US",
        "mode":       "voice",
        "version":    "1.0",
        "root":       "primary_rule",
        "tag-format": "literals",         # ⬅️  switched from semantics/1.0
    },
)

def sanitize(text: str) -> str:
    """id-safe token: lower-case, trim, spaces→_"""
    return text.lower().strip().replace(" ", "_")

def create_rule(rule_id: str, scope: str | None = None) -> ET.Element:
    rule = ET.Element("rule", id=sanitize(rule_id))
    if scope:
        rule.set("scope", scope)
    return rule

def create_item(text: str, repeat: str | None = None) -> ET.Element:
    """Explodes multi-word literals into sub-items so they can be recognized word-by-word."""
    item = ET.Element("item")
    if repeat:
        item.set("repeat", repeat)
    words = text.strip().split()
    if len(words) == 1:
        item.text = words[0]
    else:
        for w in words:
            sub = ET.Element("item")
            sub.text = w
            item.append(sub)
    return item

def create_ruleref(rule_id: str) -> ET.Element:
    return ET.Element("ruleref", uri=f"#{sanitize(rule_id)}")

def literal_tag(text: str) -> ET.Element:
    tg = ET.Element("tag")
    tg.text = text
    return tg

# ------------------------------------------------------------------ #
#  Build a map {department → [all aliases]}
# ------------------------------------------------------------------ #
aliases_by_dept: dict[str, list[str]] = {}

for col in departments_ws.iter_cols(1, departments_ws.max_column):
    if not col or not col[1].value:
        continue
    canonical = col[1].value.strip()
    aliases   = [c.value.strip().replace("&", " and ") for c in col if c.value]
    aliases_by_dept[canonical] = aliases

# ------------------------------------------------------------------ #
#  Create rules
# ------------------------------------------------------------------ #
created_depts        : set[str] = set()
created_name_only    : set[str] = set()
created_name_and_dept: set[str] = set()

primary_oneof = ET.Element("one-of")   # we’ll fill this as we go

for i in range(1, data_ws.max_row):
    if not data_ws[i][0].value:          # empty row
        continue

    dept_cell  = str(data_ws[i][3].value).strip()
    first_name = str(data_ws[i][4].value).strip()
    last_name  = str(data_ws[i][5].value).strip()
    dn         = str(data_ws[i][1].value).strip()

    canonical_name = f"{first_name} {last_name}".strip()
    dept_aliases   = aliases_by_dept.get(dept_cell, [dept_cell])

    # --------------- Department-only rule -------------------------- #
    if first_name == "Primary" and dept_cell not in created_depts:
        dept_rule = create_rule(dept_cell)
        # allow any alias
        dept_oneof = ET.Element("one-of")
        for a in dept_aliases:
            dept_oneof.append(create_item(a))
        dept_rule.append(dept_oneof)
        dept_rule.append(literal_tag(f"department {dept_cell} number {dn}"))
        grammar.append(dept_rule)

        # add to primary_rule
        pr_item = ET.Element("item")
        pr_item.append(create_ruleref(dept_cell))
        primary_oneof.append(pr_item)

        created_depts.add(dept_cell)

    # --------------- Staff member rules --------------------------- #
    elif canonical_name.lower() in staff_names:
        # -- (a) name-only ----------------------------------------- #
        if canonical_name not in created_name_only:
            name_rule = create_rule(canonical_name)
            name_rule.append(create_item(canonical_name))
            name_rule.append(literal_tag(f"{canonical_name} number {dn}"))
            grammar.append(name_rule)

            pr_item = ET.Element("item")
            pr_item.append(create_ruleref(canonical_name))
            primary_oneof.append(pr_item)

            created_name_only.add(canonical_name)

        # -- (b) name + department -------------------------------- #
        combo_id = f"{canonical_name} {dept_cell}"
        if combo_id not in created_name_and_dept:
            combo_rule = create_rule(combo_id)
            # spoken form:  John Doe ( [in] ) Sales|Customer Service|...
            combo_rule.append(create_item(canonical_name))
            combo_rule.append(create_item("in", repeat="0-1"))

            dept_item = ET.Element("item")
            dept_alias_oneof = ET.Element("one-of")
            for a in dept_aliases:
                dept_alias_oneof.append(create_item(a))
            dept_item.append(dept_alias_oneof)
            combo_rule.append(dept_item)

            combo_rule.append(
                literal_tag(f"{canonical_name} in {dept_cell} number {dn}")
            )
            grammar.append(combo_rule)

            pr_item = ET.Element("item")
            pr_item.append(create_ruleref(combo_id))
            primary_oneof.append(pr_item)

            created_name_and_dept.add(combo_id)

# ------------------------------------------------------------------ #
#  Build root rule
# ------------------------------------------------------------------ #
primary_rule = create_rule("primary_rule", scope="public")
primary_rule.append(primary_oneof)
grammar.append(primary_rule)

# ------------------------------------------------------------------ #
#  Write output
# ------------------------------------------------------------------ #
with open("Smith_Falls_lit.grxml", "wb") as fh:
    pretty = minidom.parseString(ET.tostring(grammar, encoding="utf-8"))
    fh.write(pretty.toprettyxml(indent="  ", encoding="UTF-8"))
