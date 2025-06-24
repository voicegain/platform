import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
from collections import defaultdict

# ------------------------------------------------------------------ #
#  Load spreadsheets
# ------------------------------------------------------------------ #
data_wb  = openpyxl.load_workbook("Phone-NEW customer copy May 13.xlsx")
data_ws  = data_wb.active                         # or data_wb["SheetName"]

alias_wb = openpyxl.load_workbook("departments.xlsx")
alias_ws = alias_wb.active

# ------------------------------------------------------------------ #
#  Helper builders
# ------------------------------------------------------------------ #
grammar = ET.Element(
    "grammar",
    {
        "xmlns":      "http://www.w3.org/2001/06/grammar",
        "xml:lang":   "en-US",
        "mode":       "voice",
        "version":    "1.0",
        "root":       "primary_rule",
        "tag-format": "semantics/1.0-literals",
    },
)

def sanitize(text: str) -> str:
    return text.lower().strip().replace(" ", "_")

def create_rule(rule_id: str, scope: str | None = None) -> ET.Element:
    rule = ET.Element("rule", id=sanitize(rule_id))
    if scope:
        rule.set("scope", scope)
    return rule

def create_item(text: str, repeat: str | None = None) -> ET.Element:
    itm = ET.Element("item")
    if repeat:
        itm.set("repeat", repeat)
    itm.text = text.strip()
    return itm

def create_ruleref(rule_id: str) -> ET.Element:
    return ET.Element("ruleref", uri=f"#{sanitize(rule_id)}")

def literal_tag(text: str) -> ET.Element:
    tg = ET.Element("tag"); tg.text = text
    return tg

def add_choice(parent: ET.Element, phrases: list[str], repeat: str | None = None):
    """
    Append either <item>…</item> (single) or <one-of> wrapper (multiple),
    respecting rule (2).
    """
    if len(phrases) == 1:
        parent.append(create_item(phrases[0], repeat))
    else:
        one_of = ET.Element("one-of")
        if repeat:
            one_of.set("repeat", repeat)
        for p in phrases:
            one_of.append(create_item(p))
        parent.append(one_of)

# ------------------------------------------------------------------ #
#  Build alias map  {canonical → [aliases…]}
# ------------------------------------------------------------------ #
aliases_by_dept: dict[str, list[str]] = {}
for col in alias_ws.iter_cols(1, alias_ws.max_column):
    if not col or not col[1].value:
        continue
    canonical = str(col[1].value).strip()
    aliases   = [
        str(c.value).strip().replace("&", " and ")
        for c in col
        if c.value and str(c.value).strip()
    ]
    aliases_by_dept[canonical] = aliases

# ------------------------------------------------------------------ #
#  First pass – gather info
# ------------------------------------------------------------------ #
dept_top_dn: dict[str, str] = {}          # first phone per dept
people_rows: list[tuple]   = []           # rows flagged "yes" in column G

for i in range(2, data_ws.max_row + 1):   # assume row 1 = header
    dn          = str(data_ws.cell(i, 2).value or "").strip()    # column B
    dept_cell   = str(data_ws.cell(i, 4).value or "").strip()    # column D
    first_name  = str(data_ws.cell(i, 5).value or "").strip()    # column E
    last_name   = str(data_ws.cell(i, 6).value or "").strip()    # column F
    is_person   = str(data_ws.cell(i, 7).value or "").strip().lower() == "yes"  # col G

    if dept_cell and dn and dept_cell not in dept_top_dn:
        dept_top_dn[dept_cell] = dn        # remember first phone for dept

    if is_person and first_name and last_name and dn:
        people_rows.append((first_name, last_name, dept_cell, dn))

# ------------------------------------------------------------------ #
#  Create rules
# ------------------------------------------------------------------ #
primary_oneof = ET.Element("one-of")

# --- Department-only rules ---------------------------------------- #
for dept, dn in dept_top_dn.items():
    dept_rule = create_rule(dept)
    aliases   = aliases_by_dept.get(dept, [dept])
    add_choice(dept_rule, aliases)
    dept_rule.append(literal_tag(f"department {dept} number {dn}"))
    grammar.append(dept_rule)

    pr_item = ET.Element("item"); pr_item.append(create_ruleref(dept))
    primary_oneof.append(pr_item)

# --- Person rules -------------------------------------------------- #
created_name_only    : set[str] = set()
created_name_and_dept: set[str] = set()

for first, last, dept, dn in people_rows:
    full_name = f"{first} {last}".strip()

    # (a) name-only
    if full_name not in created_name_only:
        name_rule = create_rule(full_name)
        name_rule.append(create_item(full_name))
        name_rule.append(literal_tag(f"{full_name} number {dn}"))
        grammar.append(name_rule)

        pr_item = ET.Element("item"); pr_item.append(create_ruleref(full_name))
        primary_oneof.append(pr_item)

        created_name_only.add(full_name)

    # (b) name + department
    combo_id = f"{full_name} {dept}"
    if combo_id not in created_name_and_dept:
        combo_rule = create_rule(combo_id)
        combo_rule.append(create_item(full_name))
        combo_rule.append(create_item("in", repeat="0-1"))
        aliases = aliases_by_dept.get(dept, [dept])
        add_choice(combo_rule, aliases)
        combo_rule.append(literal_tag(f"{full_name} in {dept} number {dn}"))
        grammar.append(combo_rule)

        pr_item = ET.Element("item"); pr_item.append(create_ruleref(combo_id))
        primary_oneof.append(pr_item)

        created_name_and_dept.add(combo_id)

# ------------------------------------------------------------------ #
#  Root rule
# ------------------------------------------------------------------ #
primary_rule = create_rule("primary_rule", scope="public")
primary_rule.append(primary_oneof)
grammar.append(primary_rule)

# ------------------------------------------------------------------ #
#  Write output
# ------------------------------------------------------------------ #
with open("Smith_Falls_4.grxml", "wb") as fh:
    pretty = minidom.parseString(ET.tostring(grammar, encoding="utf-8"))
    fh.write(pretty.toprettyxml(indent="  ", encoding="UTF-8"))
