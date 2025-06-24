import xml.etree.ElementTree as ET
import openpyxl
from xml.dom import minidom

departments = openpyxl.load_workbook('departments.xlsx')
departments = departments.active
data = openpyxl.load_workbook('Phone-NEW customer copy May 13.xlsx')
data = data.active

with open('staff_names.txt', 'r') as f:
    staff_names = f.read().splitlines()
    staff_names = [name.lower() for name in staff_names]

# Base of the grammar
grammar = ET.Element('grammar')
grammar.set('xmlns', 'http://www.w3.org/2001/06/grammar')
grammar.set('xml:lang', 'en-US')
grammar.set('root', 'primary_rule')
grammar.set('tag-format', 'semantics/1.0')
grammar.set('mode', 'voice')
grammar.set('version', '1.0')

# Create a rule
def create_rule(id, scope=None):
    rule = ET.Element('rule')
    rule.set('id', id.lower().strip().replace(' ', '_'))
    if scope:
        rule.set('scope', scope)
    return rule

# Create an item
def create_item(text=None, repeat=None):
    item = ET.Element('item')
    if repeat:
        item.set('repeat', repeat)
    if text:
        if (" " in text.strip()) == True:
            split_text = text.split(" ")
            for word in split_text:
                if word != "":
                    sub_item = create_item(word)
                    item.append(sub_item)
        else:
            item.text = text.strip()
    return item

# Create a ruleref
def create_ruleref(uri):
    ruleref = ET.Element('ruleref')
    ruleref.set('uri', "#" + uri.lower().strip().replace(' ', '_'))
    return ruleref

# Create a tag
def create_tag(dept, first, second, dn):
    tag = ET.Element('tag')
    if first == "Primary":
        tag.text = f"out.dept='{dept.strip()}'; out.dn='{dn}'"
    else:
        tag.text = f"out.dept='{dept.strip()}'; out.name='{first.strip()} {second.strip()}'; out.dn='{dn}'"
    tag.text = tag.text.replace('&quot;', '"')
    return tag

# Create a list of departments
created_departments = []

# Department rule
rule_dept = create_rule("department")
one_of_dept = ET.Element('one-of')

# Full name rule
rule_name = create_rule("full_name")
one_of_name = ET.Element('one-of')

# Main loop
for i in range(1, data.max_row):
    if data[i][0].value is not None:

        # if Primary = creates a department rule
        if data[i][4].value == "Primary":

            # checks if department has many names it can be recognized by
            for col in departments.iter_cols(1, departments.max_column):
                if col[1] != None and col[1].value != "":
                    for row in col:
                        if row.value == data[i][3].value and row.value not in created_departments:
                            rule = create_rule(col[0].value)
                            one_of = ET.Element('one-of')
                            for x in range(1, len(col)):
                                if col[x].value != None:
                                    item = create_item(col[x].value.replace("&", " and "))
                                    one_of.append(item)
                            rule.append(one_of)
                            tag = create_tag(col[0].value, data[i][4].value, data[i][5].value, data[i][1].value)
                            rule.append(tag)
                            grammar.append(rule)

                            item = create_item()
                            ruleref = create_ruleref(col[0].value)
                            item.append(ruleref)
                            one_of_dept.append(item)
                            for x in range(0, len(col)):
                                if col[x].value != None:
                                    created_departments.append(col[x].value)
                            break

            # Departments with only one name
            if data[i][3].value not in created_departments:
                rule = create_rule(data[i][3].value)
                item = create_item(data[i][3].value)
                rule.append(item)
                tag = create_tag(data[i][3].value, data[i][4].value, data[i][5].value, data[i][1].value)
                rule.append(tag)
                grammar.append(rule)

                item = create_item()
                ruleref = create_ruleref(data[i][3].value)
                item.append(ruleref)
                one_of_dept.append(item)
        
        # Staff names rules
        elif f"{data[i][4].value} {data[i][5].value}".lower() in staff_names:
            rule = create_rule(f"{data[i][4].value} {data[i][5].value}")
            item = create_item(f"{data[i][4].value} {data[i][5].value}")
            rule.append(item)
            tag = create_tag(data[i][3].value, data[i][4].value, data[i][5].value, data[i][1].value)
            rule.append(tag)
            grammar.append(rule)

            item = create_item()
            ruleref = create_ruleref(f"{data[i][4].value} {data[i][5].value}")
            item.append(ruleref)
            one_of_name.append(item)

rule_name.append(one_of_name)
grammar.append(rule_name)

rule_dept.append(one_of_dept)
grammar.append(rule_dept)

# Root rule
rule = create_rule(id="primary_rule", scope="public")
one_of = ET.Element('one-of')

item = create_item()
ruleref = create_ruleref("department")
item.append(ruleref)
one_of.append(item)

item = create_item()
name = create_item()
ruleref = create_ruleref("full_name")
name.append(ruleref)
item.append(name)
department = create_item(repeat="0-1")
ruleref = create_ruleref("department")
department.append(ruleref)
item.append(department)
one_of.append(item)

rule.append(one_of)
grammar.append(rule)

# Save the grammar
with open('Smith_Falls_2.grxml', 'wb') as f:
    rough_string = ET.tostring(grammar, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    f.write(reparsed.toprettyxml(indent='  ', encoding="UTF-8"))

