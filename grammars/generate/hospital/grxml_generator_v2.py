import xml.etree.ElementTree as ET
import openpyxl
from xml.dom import minidom

departments = openpyxl.load_workbook('departments.xlsx').active
data = openpyxl.load_workbook('Phone-NEW customer copy May 13.xlsx').active

with open('staff_names.txt', 'r') as f:
    staff_names = [name.lower() for name in f.read().splitlines()]

# ------------------------------------------------------------------ #
#  Helper builders
# ------------------------------------------------------------------ #
grammar = ET.Element('grammar', {
    'xmlns':        'http://www.w3.org/2001/06/grammar',
    'xml:lang':     'en-US',
    'root':         'primary_rule',
    'tag-format':   'semantics/1.0',
    'mode':         'voice',
    'version':      '1.0'
})

def create_rule(rule_id, scope=None):
    rule = ET.Element('rule', id=rule_id.lower().strip().replace(' ', '_'))
    if scope:
        rule.set('scope', scope)
    return rule

def create_item(text=None, repeat=None):
    item = ET.Element('item')
    if repeat:
        item.set('repeat', repeat)
    if text:
        # explode multi-word literals into sub-items
        for word in text.split():
            sub = ET.Element('item'); sub.text = word
            item.append(sub) if ' ' in text else None
        if ' ' not in text:
            item.text = text.strip()
    return item

def create_ruleref(uri):
    return ET.Element('ruleref', uri='#' + uri.lower().strip().replace(' ', '_'))

def create_tag(dept, first, last, dn):
    tag = ET.Element('tag')
    if first == "Primary":
        tag.text = f"out.dept='{dept.strip()}'; out.dn='{dn}'"
    else:
        tag.text = (f"out.dept='{dept.strip()}'; "
                    f"out.name='{first.strip()} {last.strip()}'; "
                    f"out.dn='{dn}'")
    return tag

# ------------------------------------------------------------------ #
#  Build department and name rules (unchanged)
# ------------------------------------------------------------------ #
created_departments = []
rule_dept = create_rule("department"); one_of_dept = ET.Element('one-of')
rule_name = create_rule("full_name");  one_of_name = ET.Element('one-of')

for i in range(1, data.max_row):
    if data[i][0].value is None:
        continue

    dept_cell   = data[i][3].value
    first_name  = data[i][4].value
    last_name   = data[i][5].value
    dn          = data[i][1].value

    # ---------------- Departments ---------------- #
    if first_name == "Primary":
        # multi-alias department
        for col in departments.iter_cols(1, departments.max_column):
            if any(col[x].value for x in range(len(col))) and col[1].value:
                if col[1].value == dept_cell and col[1].value not in created_departments:
                    rule = create_rule(col[0].value)
                    aliases = ET.Element('one-of')
                    for x in range(1, len(col)):
                        if col[x].value:
                            aliases.append(create_item(col[x].value.replace("&", " and ")))
                    rule.append(aliases)
                    rule.append(create_tag(col[0].value, first_name, last_name, dn))
                    grammar.append(rule)

                    ref_item = ET.Element('item'); ref_item.append(create_ruleref(col[0].value))
                    one_of_dept.append(ref_item)

                    created_departments += [c.value for c in col if c.value]
                    break

        # singleton department
        if dept_cell not in created_departments:
            rule = create_rule(dept_cell)
            rule.append(create_item(dept_cell))
            rule.append(create_tag(dept_cell, first_name, last_name, dn))
            grammar.append(rule)

            ref_item = ET.Element('item'); ref_item.append(create_ruleref(dept_cell))
            one_of_dept.append(ref_item)

    # ---------------- Staff (full names) ---------- #
    elif f"{first_name} {last_name}".lower() in staff_names:
        rule = create_rule(f"{first_name} {last_name}")
        rule.append(create_item(f"{first_name} {last_name}"))
        rule.append(create_tag(dept_cell, first_name, last_name, dn))
        grammar.append(rule)

        ref_item = ET.Element('item'); ref_item.append(create_ruleref(f"{first_name} {last_name}"))
        one_of_name.append(ref_item)

rule_name.append(one_of_name); grammar.append(rule_name)
rule_dept.append(one_of_dept); grammar.append(rule_dept)

# ------------------------------------------------------------------ #
#  NEW: primary_rule with explicit tag propagation
# ------------------------------------------------------------------ #
primary = create_rule('primary_rule', scope='public')
one_of  = ET.Element('one-of')

# (1) department only ------------------------------------------------
item_dept = ET.Element('item')
item_dept.append(create_ruleref('department'))
tag_dept  = ET.Element('tag'); tag_dept.text = "out = $department;"
item_dept.append(tag_dept)
one_of.append(item_dept)

# (2) full_name only --------------------------------------------------
item_name = ET.Element('item')
item_name.append(create_ruleref('full_name'))
tag_name  = ET.Element('tag'); tag_name.text = "out = $full_name;"
item_name.append(tag_name)
one_of.append(item_name)

# (3) full_name + department -----------------------------------------
item_both = ET.Element('item')
item_both.append(create_ruleref('full_name'))
item_both.append(create_ruleref('department'))
tag_both  = ET.Element('tag'); tag_both.text = (
    "out = $full_name; "          # start with the person object
    "out.dept = $department.dept; "
    "out.dn   = $department.dn;"  # pull fields you need
)
item_both.append(tag_both)
one_of.append(item_both)

primary.append(one_of)
grammar.append(primary)

# ------------------------------------------------------------------ #
#  Write the grammar file
# ------------------------------------------------------------------ #
with open('Smith_Falls_2.grxml', 'wb') as f:
    pretty = minidom.parseString(ET.tostring(grammar, encoding='utf-8'))
    f.write(pretty.toprettyxml(indent='  ', encoding='UTF-8'))
